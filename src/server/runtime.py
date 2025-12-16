from __future__ import annotations

import asyncio
import json
import re
import time
from collections.abc import Iterable
from pathlib import Path
from types import GeneratorType
from typing import Any, Dict, List, Optional

import duckdb

from ..agent import (
    DirectOpenAIAgent,
    DuckDBComponents,
    DuckDBWorkspaceConfig,
    create_duckdb_agent,
    create_duckdb_components,
    resolve_duckdb_workspace,
)
from ..tools.datalake.manager import DatalakeManager
from ..tools.duckdb_query_tool import DuckDBQueryTool, StructuredQueryResult
from ..tools.duckdb_schema_manager import SchemaManager
from .chat_history import ChatHistoryManager
from .config import ServerSettings
from .models import (
    AIAssistFieldResponse,
    ChatMessage,
    DatabaseStatsResponse,
    ReclaimSpaceResponse,
    TableStats,
)


class DuckDBRuntime:
    """Holds shared state for CLI-compatible DuckDB tooling and the FastAPI server."""

    def __init__(self, *, settings: ServerSettings | None = None) -> None:
        # Initialize runtime with settings (reloaded)
        self.settings = settings or ServerSettings.load()
        self.workspace: DuckDBWorkspaceConfig = resolve_duckdb_workspace()
        self.components: DuckDBComponents = create_duckdb_components(self.workspace, emit_status=False)

        self._query_lock = asyncio.Lock()
        self._agent_init_lock = asyncio.Lock()
        self._agent_run_lock = asyncio.Lock()
        self._agent: Optional[DirectOpenAIAgent] = None

        # Initialize chat history manager
        self.chat_history_manager = ChatHistoryManager(self.settings)

        # Initialize datalake manager if datalakes are configured
        self.datalake_manager: Optional[DatalakeManager] = None
        if self.settings.datalakes:
            datalake_configs = [
                {
                    "name": dl.name,
                    "type": dl.type,
                    "connection_string": dl.connection_string,
                    "container_name": dl.container_name,
                }
                for dl in self.settings.datalakes
            ]
            from ..tools.datalake.manager import DatalakeConfig as DLConfig
            from ..tools.datalake.manager import DatalakeManager as DLManager

            self.datalake_manager = DLManager([
                DLConfig(**config) for config in datalake_configs
            ])

    @property
    def query_tool(self) -> DuckDBQueryTool:
        return self.components.query_tool

    @property
    def schema_manager(self) -> SchemaManager:
        return self.query_tool.schema_manager

    async def list_projects(self) -> List[Dict[str, Any]]:
        async with self._query_lock:
            tool = self.query_tool
            available = tool.discover_databases()
            current = tool.current_database
            projects: List[Dict[str, Any]] = []
            for path in available:
                metadata = self._project_metadata_from_path(path)
                metadata["is_active"] = bool(current and path == current)
                projects.append(metadata)
            return projects

    async def _ensure_project_locked(self, *, project: Optional[str], database: Optional[str]) -> Optional[str]:
        tool = self.query_tool
        available = tool.discover_databases()
        if not available:
            return None
        target = tool._resolve_database(database, project, available)
        tool.current_database = target
        tool.schema_manager.set_database(target)
        return str(target)

    async def get_schema(self, *, project: Optional[str] = None, database: Optional[str] = None) -> Dict[str, Any]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                return {}
            return self.query_tool.schema_manager.schema

    async def list_tables(self, *, project: Optional[str] = None, database: Optional[str] = None) -> List[str]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                return []
            try:
                tables = await asyncio.to_thread(self.query_tool.list_tables_in_database)
            except ValueError:
                tables = self.query_tool.schema_manager.list_tables()

            if not tables:
                tables = self.query_tool.schema_manager.list_tables()
            return tables

    async def list_diagrams(
        self,
        *,
        project: Optional[str] = None,
        database: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                return []
            manager = self.schema_manager
            return manager.list_diagrams()

    async def save_diagram(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        name: str,
        description: Optional[str],
        tables: List[Dict[str, Any]],
        diagram_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to save diagrams.")
            manager = self.schema_manager
            return manager.save_diagram(name=name, description=description, tables=tables, diagram_id=diagram_id)

    async def delete_diagram(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        diagram_id: str,
    ) -> None:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to delete diagrams.")
            manager = self.schema_manager
            removed = manager.delete_diagram(diagram_id)
            if not removed:
                raise ValueError("Diagram not found.")

    async def list_queries(
        self,
        *,
        project: Optional[str] = None,
        database: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                return []
            manager = self.schema_manager
            return manager.list_queries()

    async def save_query(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        name: str,
        description: Optional[str],
        sql: str,
        limit: Optional[int],
        query_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to save queries.")
            manager = self.schema_manager
            return manager.save_query(
                name=name,
                description=description,
                sql=sql,
                limit=limit,
                query_id=query_id,
            )

    async def delete_query(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        query_id: str,
    ) -> None:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to delete queries.")
            manager = self.schema_manager
            removed = manager.delete_query(query_id)
            if not removed:
                raise ValueError("Query not found.")

    async def list_undocumented_fields(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
    ) -> List[Dict[str, Any]]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                return []
            manager = self.schema_manager
            table_record = manager.get_table(table) or {}
            documented_fields = table_record.get("fields", {}) or {}
            documented_lower = {name.lower() for name in documented_fields.keys()}

            try:
                schema_rows = await asyncio.to_thread(self.query_tool.fetch_table_schema, table)
            except ValueError:
                return []

            missing: list[Dict[str, Any]] = []
            for row in schema_rows:
                column_name = str(row.get("name") or row.get("column_name") or "")
                if not column_name:
                    continue
                if column_name.lower() not in documented_lower:
                    data_type = str(row.get("type") or row.get("column_type") or "").strip()
                    missing.append(
                        {
                            "name": column_name,
                            "data_type": data_type or None,
                        }
                    )

            missing.sort(key=lambda value: str(value.get("name", "")).lower())
            return missing

    async def update_table_metadata(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
        short_description: Optional[str],
        long_description: Optional[str],
    ) -> Dict[str, Any]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to update schema.")
            manager = self.schema_manager
            metadata = manager.update_table(
                table,
                short_description=short_description,
                long_description=long_description,
            )
            return {"table": table, "metadata": metadata}

    async def delete_table(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
    ) -> None:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to delete table.")
            manager = self.schema_manager
            removed = manager.delete_table(table)
            if not removed:
                raise ValueError(f"Table '{table}' not found in schema.")

    async def delete_table_data(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
        mode: str,
        keep_count: Optional[int] = None,
        sort_column: Optional[str] = None,
        sort_order: Optional[str] = None,
        filter_condition: Optional[str] = None,
    ) -> Dict[str, Any]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to delete data.")

            if mode == "all":
                sql = f'DELETE FROM "{table}"'
                message = f"Deleted all data from table '{table}'"
            elif mode == "subset":
                if not keep_count or not sort_column or not sort_order:
                    raise ValueError("Missing parameters for subset deletion")

                # Use rowid to identify rows to keep
                # We want to keep the top N rows ordered by sort_column
                # So we delete rows whose rowid is NOT in that set
                sql = f"""
                    DELETE FROM "{table}"
                    WHERE rowid NOT IN (
                        SELECT rowid
                        FROM "{table}"
                        ORDER BY "{sort_column}" {sort_order}
                        LIMIT {keep_count}
                    )
                """
                message = f"Retained top {keep_count} rows in '{table}' ordered by {sort_column} {sort_order}"
            elif mode == "filter":
                if not filter_condition:
                    raise ValueError("Missing filter condition for filter deletion")

                # Delete rows NOT matching the filter (Retain mode)
                # Using IS NOT TRUE to handle NULLs correctly
                # (NULLs should be deleted if they don't match the keep condition)
                sql = f'DELETE FROM "{table}" WHERE ({filter_condition}) IS NOT TRUE'
                message = f"Retained rows in '{table}' matching filter: {filter_condition}"
            else:
                raise ValueError(f"Invalid deletion mode: {mode}")

            await asyncio.to_thread(self.query_tool.execute_structured, sql)
            return {"success": True, "message": message}

    def _format_size(self, size_bytes: Optional[int]) -> str:
        if size_bytes is None:
            return "N/A"

        current_size = float(size_bytes)
        for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
            if current_size < 1024.0:
                return f"{current_size:.2f} {unit}"
            current_size /= 1024.0
        return f"{current_size:.2f} PB"

    async def get_database_stats(
        self,
        *,
        project: Optional[str] = None,
        database: Optional[str] = None,
    ) -> DatabaseStatsResponse:
        print(f"DEBUG: get_database_stats called for project={project}, database={database}")
        try:
            async with self._query_lock:
                resolved = await self._ensure_project_locked(project=project, database=database)
                if resolved is None:
                    print("DEBUG: No database resolved")
                    return DatabaseStatsResponse(tables=[])

                # Calculate total file size
                total_size_bytes = None
                total_size_pretty = None
                try:
                    db_path = Path(resolved)
                    if db_path.exists():
                        total_size_bytes = db_path.stat().st_size
                        total_size_pretty = self._format_size(total_size_bytes)
                except Exception as e:
                    print(f"DEBUG: Error getting file size: {e}")

                # 1. Get all tables
                # Avoid calling self.list_tables() to prevent deadlock (re-entrant lock)
                try:
                    tables = await asyncio.to_thread(self.query_tool.list_tables_in_database)
                except ValueError:
                    tables = self.query_tool.schema_manager.list_tables()

                if not tables:
                    tables = self.query_tool.schema_manager.list_tables()

                print(f"DEBUG: Found {len(tables)} tables")

                if not tables:
                    return DatabaseStatsResponse(
                        tables=[],
                        total_size_bytes=total_size_bytes,
                        total_size_pretty=total_size_pretty
                    )

                stats = []
                for table in tables:
                    # Handle quoting
                    if "." in table:
                        schema, name = table.split(".", 1)
                        safe_schema = schema.replace('"', '""')
                        safe_name = name.replace('"', '""')
                        escaped = f'"{safe_schema}"."{safe_name}"'
                    else:
                        safe_table = table.replace('"', '""')
                        escaped = f'"{safe_table}"'

                    try:
                        # Simple COUNT(*) per table
                        res = await asyncio.to_thread(
                            self.query_tool.execute_structured,
                            f"SELECT COUNT(*) FROM {escaped}"
                        )
                        count = int(res.rows[0][0]) if res.rows else 0
                    except Exception as e:
                        print(f"DEBUG: Error counting {table}: {e}")
                        count = 0

                    stats.append(TableStats(
                        name=table,
                        row_count=count,
                        size_bytes=None,
                        size_pretty="N/A"
                    ))

                return DatabaseStatsResponse(
                    tables=stats,
                    total_size_bytes=total_size_bytes,
                    total_size_pretty=total_size_pretty
                )
        except Exception as e:
            print(f"ERROR in get_database_stats: {e}")
            import traceback
            traceback.print_exc()
            raise

    async def reclaim_space(
        self,
        *,
        project: Optional[str] = None,
        database: Optional[str] = None,
    ) -> ReclaimSpaceResponse:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to reclaim space.")

            db_path = Path(resolved)
            old_size = db_path.stat().st_size if db_path.exists() else 0

            # Execute VACUUM
            await asyncio.to_thread(self.query_tool.execute_structured, "VACUUM")

            new_size = db_path.stat().st_size if db_path.exists() else 0
            freed_bytes = max(0, old_size - new_size)
            freed_pretty = self._format_size(freed_bytes)

            return ReclaimSpaceResponse(
                success=True,
                message="Space reclaimed successfully",
                old_size_bytes=old_size,
                new_size_bytes=new_size,
                freed_bytes=freed_bytes,
                freed_pretty=freed_pretty
            )

    async def update_field_metadata(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
        field: str,
        short_description: Optional[str],
        long_description: Optional[str],
        nullability: Optional[str],
        data_type: Optional[str],
        values: Optional[Dict[str, str]],
        relationships: Optional[List[Dict[str, Any]]],
        new_field_name: Optional[str],
        allow_null: Optional[bool],
        ignored: Optional[bool] = None,
    ) -> Dict[str, Any]:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to update schema.")
            manager = self.schema_manager
            nullability_text: Optional[str]
            if allow_null is None:
                nullability_text = nullability
            else:
                nullability_text = "nullable" if allow_null else "not nullable"
            metadata = manager.update_field(
                table,
                field,
                short_description=short_description,
                long_description=long_description,
                nullability=nullability_text,
                data_type=data_type,
                values=values,
                ignored=ignored,
            )
            updated_field_name = field
            if new_field_name and new_field_name.strip() and new_field_name.strip() != field:
                metadata = manager.rename_field(table, field, new_field_name.strip())
                updated_field_name = new_field_name.strip()

            existing_table = manager.get_table(table) or {}
            existing_relationships = list(existing_table.get("relationships") or [])

            if relationships is not None:
                normalized: list[dict[str, object]] = []
                target_fields = {field.lower(), updated_field_name.lower()}

                for rel in existing_relationships:
                    rel_field = str(rel.get("field") or "").lower()
                    if rel_field not in target_fields:
                        normalized.append(dict(rel))

                for rel in relationships:
                    related_table = rel.get("related_table")
                    related_field = rel.get("related_field")
                    if not related_table or not related_field:
                        continue
                    normalized.append(
                        {
                            "field": updated_field_name,
                            "related_table": related_table,
                            "related_field": related_field,
                            "type": rel.get("type") or "unspecified",
                        }
                    )
                manager.set_relationships(table, normalized)
                metadata = manager.get_field(table, updated_field_name) or metadata
            return {"field": updated_field_name, "metadata": metadata}

    async def run_sql(
        self,
        *,
        sql: str,
        project: Optional[str] = None,
        database: Optional[str] = None,
        limit: Optional[int] = None,
    ) -> StructuredQueryResult:
        async with self._query_lock:
            result_limit = limit if limit is not None else self.settings.query_default_limit
            return await asyncio.to_thread(
                self.query_tool.execute_structured,
                sql,
                database=database,
                project=project,
                result_limit=result_limit,
            )

    async def auto_describe_field(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
        field: str,
        current_short_description: Optional[str] = None,
        current_long_description: Optional[str] = None,
        data_type: Optional[str] = None,
        description_type: Optional[str] = None,
    ) -> str:
        await self._ensure_project_locked(project=project, database=database)
        agent = await self._ensure_agent()
        # Set agent's project context to avoid "Multiple DuckDB projects detected" error
        self._set_agent_project(agent, project=project, database=database)

        manager = self.schema_manager
        table_record = manager.get_table(table) or {}
        field_record = manager.get_field(table, field) or {}

        is_short = description_type == 'short'

        if description_type == 'data_type':
            # 1. Check actual schema
            try:
                schema_rows = await asyncio.to_thread(self.query_tool.fetch_table_schema, table)
                for row in schema_rows:
                    if str(row.get('name', '')).lower() == field.lower():
                        return str(row.get('type', '')).upper()
            except Exception:
                pass

            # 2. If not found in DB, try to infer from metadata values if available
            values = field_record.get("values", {})
            if values:
                prompt = (
                    f"Infer the DuckDB data type for field '{field}' in table '{table}'. "
                    f"It has these allowed values: {', '.join(values.keys())}. "
                    "Return ONLY the data type (e.g. VARCHAR, INTEGER, BOOLEAN)."
                )
                try:
                    response = await asyncio.to_thread(agent.run, prompt, reset=False)
                    return self._materialize_agent_response(response).strip()
                except Exception:
                    pass

            return "UNKNOWN"

        if is_short:
            prompt_lines = [
                "You are assisting with documenting DuckDB schemas. "
                "Write a brief short description for the given field.",
                "Keep it very concise (5-10 words maximum), use plain text, and avoid Markdown formatting.",
                "The description should be a brief phrase or label that summarizes what this field contains.",
                "",
                f"Table: {table}",
                f"Field: {field}",
                f"Data type: {data_type or field_record.get('data_type') or 'unknown'}",
            ]
        else:
            prompt_lines = [
                "You are assisting with documenting DuckDB schemas. "
                "Write a rich long description for the given field.",
                "Keep it concise (2-3 sentences), use plain text, and avoid Markdown formatting.",
                "",
                f"Table: {table}",
                f"Field: {field}",
                f"Data type: {data_type or field_record.get('data_type') or 'unknown'}",
            ]
        if current_short_description or field_record.get("short_description"):
            prompt_lines.append(
                f"Short description: {current_short_description or field_record.get('short_description')}"
            )
        if table_record.get("short_description") or table_record.get("long_description"):
            prompt_lines.append(
                "Table context: "
                + " ".join(
                    filter(
                        None,
                        [
                            str(table_record.get("short_description") or ""),
                            str(table_record.get("long_description") or ""),
                        ],
                    )
                )
            )

        prompt = "\n".join(line for line in prompt_lines if line)

        try:
            response = await asyncio.to_thread(agent.run, prompt, reset=False)
            return self._materialize_agent_response(response).strip()
        except Exception as e:
            import traceback
            print(f"Error in auto_describe_field: {e}")
            print(f"Full traceback:\n{traceback.format_exc()}")
            raise

    async def ai_assist_field(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
        field: str,
    ) -> "AIAssistFieldResponse":
        """
        AI assist for field: fetches sample values from the database and generates
        short description, long description, data type, and nullable status.
        """
        await self._ensure_project_locked(project=project, database=database)
        agent = await self._ensure_agent()
        # Set agent's project context
        self._set_agent_project(agent, project=project, database=database)

        # Fetch sample values and null count from the database
        sample_values = []
        has_nulls = False
        try:
            sql = f'SELECT DISTINCT "{field}" FROM "{table}" WHERE "{field}" IS NOT NULL LIMIT 20'
            result = await asyncio.to_thread(self.query_tool.execute_structured, sql)
            if result and hasattr(result, 'rows') and result.rows:
                sample_values = [str(row[0]) for row in result.rows if row]

            # Check for null values
            null_sql = f'SELECT COUNT(*) FROM "{table}" WHERE "{field}" IS NULL'
            null_result = await asyncio.to_thread(self.query_tool.execute_structured, null_sql)
            if null_result and hasattr(null_result, 'rows') and null_result.rows:
                null_count = null_result.rows[0][0]
                has_nulls = null_count > 0
        except Exception as e:
            print(f"Warning: Could not fetch sample values for {table}.{field}: {e}")

        # Get table context
        manager = self.schema_manager
        table_record = manager.get_table(table) or {}
        table_context = " ".join(
            filter(
                None,
                [
                    str(table_record.get("short_description") or ""),
                    str(table_record.get("long_description") or ""),
                ],
            )
        )

        # Build prompt for AI
        prompt_parts = [
            "You are assisting with documenting DuckDB database schemas.",
            f"Analyze the field '{field}' in table '{table}'.",
        ]

        if table_context:
            prompt_parts.append(f"Table context: {table_context}")

        if sample_values:
            sample_str = ", ".join(sample_values[:15])
            if len(sample_values) > 15:
                sample_str += f"... (and {len(sample_values) - 15} more)"
            prompt_parts.append(f"Sample values from the database: {sample_str}")

        prompt_parts.append(f"Null values found: {'Yes' if has_nulls else 'No'}")

        prompt_parts.extend([
            "",
            "Based on this information, provide:",
            "1. A short description (5-10 words, concise label)",
            "2. A long description (2-3 sentences, detailed explanation)",
            "3. The DuckDB data type - Important guidelines:",
            "   - If values match pattern XXXXXXXX-XXXX-XXXX-XXXX-XXXXXXXXXXXX (8-4-4-4-12 hex characters), use UUID",
            "   - Common types: VARCHAR, INTEGER, BOOLEAN, DATE, TIMESTAMP, UUID, DECIMAL",
            "   - Choose the most specific type that matches the data pattern",
            "4. Whether the field is nullable (true/false) - if no null values were found, assume NOT NULL",
            "",
            "Format your response EXACTLY as follows (use these exact labels):",
            "SHORT: <short description>",
            "LONG: <long description>",
            "TYPE: <data type>",
            "NULLABLE: <true or false>",
            "",
            "Use plain text only, no markdown formatting.",
        ])

        prompt = "\n".join(prompt_parts)

        try:
            response = await asyncio.to_thread(agent.run, prompt, reset=False)
            text_response = self._materialize_agent_response(response).strip()

            # Parse the response
            short_desc = ""
            long_desc = ""
            data_type = "VARCHAR"
            nullable = True

            for line in text_response.split('\n'):
                line = line.strip()
                if line.startswith("SHORT:"):
                    short_desc = line[6:].strip()
                elif line.startswith("LONG:"):
                    long_desc = line[5:].strip()
                elif line.startswith("TYPE:"):
                    data_type = line[5:].strip()
                elif line.startswith("NULLABLE:"):
                    val = line[9:].strip().lower()
                    nullable = val == 'true'

            return AIAssistFieldResponse(
                short_description=short_desc,
                long_description=long_desc,
                data_type=data_type,
                nullable=nullable,
            )
        except Exception as e:
            import traceback
            print(f"Error in ai_assist_field: {e}")
            print(f"Full traceback:\n{traceback.format_exc()}")
            raise

    async def _ensure_agent(self) -> DirectOpenAIAgent:
        if self._agent is not None:
            return self._agent

        async with self._agent_init_lock:
            if self._agent is None:
                self._agent = create_duckdb_agent(
                    search_roots=self.workspace.search_roots,
                    default_database=self.workspace.default_database,
                    duckdb_executable=self.workspace.duckdb_executable,
                    max_steps=self.settings.agent_max_steps,
                    timeout=self.settings.agent_timeout,
                    max_retries=self.settings.agent_max_retries,
                    emit_status=False,
                )
        return self._agent

    def _set_agent_project(
        self,
        agent: DirectOpenAIAgent,
        *,
        project: Optional[str],
        database: Optional[str],
    ) -> Optional[str]:
        tool = getattr(agent, "duckdb_tool", None)
        if not isinstance(tool, DuckDBQueryTool):
            return None
        available = tool.discover_databases()
        if not available:
            return None
        target = tool._resolve_database(database, project, available)
        tool.current_database = target
        tool.schema_manager.set_database(target)

        # Also set current_database for chart_tool
        chart_tool = getattr(agent, "duckdb_chart_tool", None)
        if chart_tool is not None:
            chart_tool.current_database = str(target)

        return str(target)

    async def run_chat(
        self,
        *,
        message: str,
        reset: bool = False,
        project: Optional[str] = None,
        database: Optional[str] = None,
        file_content: Optional[str] = None,
        filename: Optional[str] = None,
        session_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Run a chat interaction with the AI agent.
        """
        agent = await self._ensure_agent()
        async with self._agent_run_lock:
            active_db = self._set_agent_project(agent, project=project, database=database)

            # Restore history if session exists
            if session_id and project and not reset:
                session = self.chat_history_manager.get_session(project, session_id)
                if session:
                    restored_history = []
                    for msg in session.messages:
                        msg_dict: Dict[str, Any] = {"role": msg.role, "content": msg.content or ""}
                        if msg.tool_calls:
                            msg_dict["tool_calls"] = msg.tool_calls
                        if msg.tool_call_id:
                            msg_dict["tool_call_id"] = msg.tool_call_id
                        restored_history.append(msg_dict)
                    agent.conversation_history = self._sanitize_history(restored_history)

            enhanced_message = message

            # Append file content if provided
            if file_content and filename:
                processed_content = self._process_file_content(filename, file_content)
                enhanced_message += f"\n\n[User uploaded file: {filename}]\n{processed_content}"

            # Prepend project context if available
            if active_db:
                project_name = Path(active_db).stem
                db_filename = Path(active_db).name
                enhanced_message = (
                    f"[Context: You are currently working with the '{project_name}' DuckDB project. "
                    f"The database '{db_filename}' is already selected and active. "
                    f"You do not need to specify database or project parameters in tool calls.]\n\n"
                    f"{enhanced_message}"
                )

            response = await asyncio.to_thread(agent.run, enhanced_message, reset=reset)
            reply_text = self._materialize_agent_response(response)

            # Save session
            current_session_id = session_id
            if project:
                messages = []
                for msg in agent.conversation_history:
                    # Skip system messages
                    if msg.get("role") == "system":
                        continue

                    content = msg.get("content") or ""

                    # Strip context from user messages (only if content is a string)
                    if msg.get("role") == "user" and isinstance(content, str) and content.startswith("[Context:"):
                        # Find the end of the context block
                        context_end = content.find("]\n\n")
                        if context_end != -1:
                            content = content[context_end + 3:]
                        else:
                            # Fallback for different formatting
                            content = re.sub(r"^\[Context:.*?\]\n\n", "", content, flags=re.DOTALL)

                    # For multimodal messages (list content), convert to JSON string for storage
                    if isinstance(content, list):
                        import json
                        content = json.dumps(content)

                    messages.append(ChatMessage(
                        role=str(msg.get("role") or "user"),
                        content=content,
                        timestamp=time.time(),
                        tool_calls=self._serialize_tool_calls(msg.get("tool_calls")),
                        tool_call_id=msg.get("tool_call_id")
                    ))

                saved_session = self.chat_history_manager.save_session(project, messages, session_id)
                current_session_id = saved_session.id

            return {
                "reply": reply_text,
                "database": active_db,
                "metadata": getattr(agent, 'execution_metadata', {}),
                "session_id": current_session_id,
            }

    async def run_chat_stream(
        self,
        *,
        message: str,
        reset: bool = False,
        project: Optional[str] = None,
        database: Optional[str] = None,
        file_content: Optional[str] = None,
        filename: Optional[str] = None,
        session_id: Optional[str] = None,
        images: Optional[List[Dict[str, str]]] = None,
    ):
        """Stream chat responses as they are generated."""
        print(f"DEBUG: run_chat_stream called with message={message}, project={project}")
        agent = await self._ensure_agent()
        print("DEBUG: agent initialized")
        async with self._agent_run_lock:
            print("DEBUG: acquired agent_run_lock")
            active_db = self._set_agent_project(agent, project=project, database=database)
            print(f"DEBUG: set agent project, active_db={active_db}")

            # Restore history if session exists
            if session_id and project and not reset:
                session = self.chat_history_manager.get_session(project, session_id)
                if session:
                    restored_history = []
                    for msg in session.messages:
                        msg_dict: Dict[str, Any] = {"role": msg.role, "content": msg.content or ""}
                        if msg.tool_calls:
                            msg_dict["tool_calls"] = msg.tool_calls
                        if msg.tool_call_id:
                            msg_dict["tool_call_id"] = msg.tool_call_id
                        restored_history.append(msg_dict)
                    agent.conversation_history = self._sanitize_history(restored_history)

            enhanced_message = message

            # Append file content if provided
            if file_content and filename:
                processed_content = self._process_file_content(filename, file_content)
                enhanced_message += f"\n\n[User uploaded file: {filename}]\n{processed_content}"

            # Prepend project context to the message
            if active_db:
                project_name = Path(active_db).stem
                db_filename = Path(active_db).name

                # Get query instructions
                query_instructions = ""
                try:
                    # Access schema manager from the tool attached to the agent
                    tool = getattr(agent, "duckdb_tool", None)
                    if tool and hasattr(tool, "schema_manager"):
                        meta = tool.schema_manager.get_project_metadata()
                        query_instructions = meta.get("query_instructions", "")
                except Exception:
                    pass

                instructions_text = ""
                if query_instructions:
                    instructions_text = f"\n[Project Instructions: {query_instructions}]\n"

                enhanced_message = (
                    f"[Context: You are currently working with the '{project_name}' DuckDB project. "
                    f"The database '{db_filename}' is already selected and active. "
                    f"You do not need to specify database or project parameters in tool calls.]{instructions_text}\n\n"
                    f"{enhanced_message}"
                )
            else:
                enhanced_message = message

            try:
                # Setup queue for streaming status updates
                status_queue = asyncio.Queue()
                loop = asyncio.get_running_loop()

                def on_step_callback(msg: str):
                    loop.call_soon_threadsafe(status_queue.put_nowait, {"type": "status", "message": msg})

                # Run agent in thread and capture response with timeout
                print("DEBUG: calling agent.run in thread")
                # Allow enough time for multiple steps with the configured timeout
                # Increased to 300s to handle potential rate limit retries
                execution_timeout = max(300.0, self.settings.agent_timeout * 5)

                agent_task = asyncio.create_task(
                    asyncio.to_thread(
                        agent.run,
                        enhanced_message,
                        reset=reset,
                        images=images,
                        on_step=on_step_callback
                    )
                )

                # Poll queue while agent is running
                queue_task = asyncio.create_task(status_queue.get())

                while not agent_task.done():
                    try:
                        done, pending = await asyncio.wait(
                            [queue_task, agent_task],
                            return_when=asyncio.FIRST_COMPLETED
                        )

                        if queue_task in done:
                            item = queue_task.result()
                            yield item
                            # Create new task for next item
                            queue_task = asyncio.create_task(status_queue.get())

                        # If agent_task is done, the loop condition will handle it
                    except Exception:
                        pass

                # Cancel the pending queue wait
                if not queue_task.done():
                    queue_task.cancel()
                    try:
                        await queue_task
                    except asyncio.CancelledError:
                        pass

                # Process any remaining items in queue (that might have been put just before agent finished)
                while not status_queue.empty():
                    yield await status_queue.get()

                # Get the final result (or raise exception if failed)
                try:
                    response = await agent_task
                except Exception as e:
                    # Check if this is an invalid request error due to incomplete tool calls
                    error_str = str(e).lower()
                    if "toolcalls" in error_str or "tool_calls" in error_str or "invalidrequest" in error_str:
                        print(f"DEBUG: Detected corrupted conversation state, resetting: {e}")
                        # Reset the conversation and retry
                        # Note: We don't stream status for the retry to keep it simple for now
                        response = await asyncio.wait_for(
                            asyncio.to_thread(agent.run, enhanced_message, reset=True, images=images),
                            timeout=execution_timeout
                        )
                        print("DEBUG: Successfully recovered from corrupted state")
                    else:
                        # Re-raise other exceptions
                        raise

                print(f"DEBUG: agent.run completed, response type={type(response)}")
                reply_text = self._materialize_agent_response(response)
                print(f"DEBUG: materialized response, length={len(reply_text)}")
                print(f"DEBUG: reply_text content:\n{reply_text[:500]}...")  # First 500 chars

                # Save session
                current_session_id = session_id
                if project:
                    messages = []
                    for msg in agent.conversation_history:
                        # Skip system messages
                        if msg.get("role") == "system":
                            continue

                        content = msg.get("content") or ""

                        # Strip context from user messages (only if content is a string)
                        if msg.get("role") == "user" and isinstance(content, str) and content.startswith("[Context:"):
                            context_end = content.find("]\n\n")
                            if context_end != -1:
                                content = content[context_end + 3:]
                            else:
                                content = re.sub(r"^\[Context:.*?\]\n\n", "", content, flags=re.DOTALL)

                        # For multimodal messages (list content), convert to JSON string for storage
                        if isinstance(content, list):
                            import json
                            content = json.dumps(content)

                        messages.append(ChatMessage(
                            role=str(msg.get("role") or "user"),
                            content=content,
                            timestamp=time.time(),
                            tool_calls=self._serialize_tool_calls(msg.get("tool_calls")),
                            tool_call_id=msg.get("tool_call_id")
                        ))

                    saved_session = self.chat_history_manager.save_session(project, messages, session_id)
                    current_session_id = saved_session.id

                # Send the complete response
                yield {
                    "type": "response",
                    "reply": reply_text,
                    "database": active_db,
                    "metadata": getattr(agent, 'execution_metadata', {}),
                    "session_id": current_session_id,
                }
                print("DEBUG: sent response message")

                # Send completion signal
                yield {"type": "done", "database": active_db, "session_id": current_session_id}
                print("DEBUG: sent done message")

            except asyncio.TimeoutError:
                print("DEBUG: agent execution timed out")
                error_msg = "Request timed out after 2 minutes. The query may be too complex or the database too large."

                # Append partial results if available
                if agent._last_implementation_plan:
                    error_msg += f"\n\n[PLAN:{agent._last_implementation_plan}]"
                if agent._last_sql_query:
                    error_msg += f"\n\n```sql\n{agent._last_sql_query}\n```"

                # Save session with error
                if project and session_id:
                    try:
                        agent.conversation_history.append({
                            "role": "assistant",
                            "content": f"Error: {error_msg}"
                        })
                        messages = []
                        for msg in agent.conversation_history:
                            if msg.get("role") == "system":
                                continue
                            content = msg.get("content") or ""
                            is_user_msg = msg.get("role") == "user"
                            if is_user_msg and isinstance(content, str) and content.startswith("[Context:"):
                                context_end = content.find("]\n\n")
                                if context_end != -1:
                                    content = content[context_end + 3:]
                                else:
                                    content = re.sub(r"^\[Context:.*?\]\n\n", "", content, flags=re.DOTALL)
                            if isinstance(content, list):
                                import json
                                content = json.dumps(content)
                            messages.append(ChatMessage(
                                role=str(msg.get("role") or "user"),
                                content=content,
                                timestamp=time.time(),
                                tool_calls=self._serialize_tool_calls(msg.get("tool_calls")),
                                tool_call_id=msg.get("tool_call_id")
                            ))
                        self.chat_history_manager.save_session(project, messages, session_id)
                    except Exception as save_err:
                        print(f"Failed to save session on timeout: {save_err}")

                yield {
                    "type": "error",
                    "error": error_msg,
                    "database": active_db,
                }
            except Exception as e:
                print(f"DEBUG: exception caught: {e}")
                import traceback
                traceback.print_exc()

                error_msg = str(e)

                # Append partial results if available
                if agent._last_implementation_plan:
                    error_msg += f"\n\n[PLAN:{agent._last_implementation_plan}]"
                if agent._last_sql_query:
                    error_msg += f"\n\n```sql\n{agent._last_sql_query}\n```"

                # Save session with error
                if project and session_id:
                    try:
                        agent.conversation_history.append({
                            "role": "assistant",
                            "content": f"Error: {error_msg}"
                        })
                        messages = []
                        for msg in agent.conversation_history:
                            if msg.get("role") == "system":
                                continue
                            content = msg.get("content") or ""
                            is_user_msg = msg.get("role") == "user"
                            if is_user_msg and isinstance(content, str) and content.startswith("[Context:"):
                                context_end = content.find("]\n\n")
                                if context_end != -1:
                                    content = content[context_end + 3:]
                                else:
                                    content = re.sub(r"^\[Context:.*?\]\n\n", "", content, flags=re.DOTALL)
                            if isinstance(content, list):
                                import json
                                content = json.dumps(content)
                            messages.append(ChatMessage(
                                role=str(msg.get("role") or "user"),
                                content=content,
                                timestamp=time.time(),
                                tool_calls=self._serialize_tool_calls(msg.get("tool_calls")),
                                tool_call_id=msg.get("tool_call_id")
                            ))
                        self.chat_history_manager.save_session(project, messages, session_id)
                    except Exception as save_err:
                        print(f"Failed to save session on error: {save_err}")

                yield {
                    "type": "error",
                    "error": error_msg,
                    "database": active_db,
                }

    async def assist_query(
        self,
        *,
        sql: str,
        project: Optional[str] = None,
        database: Optional[str] = None,
    ) -> str:
        """
        Use the AI agent to improve or fix a SQL query.
        """
        agent = await self._ensure_agent()
        async with self._agent_run_lock:
            active_db = self._set_agent_project(agent, project=project, database=database)

            prompt = (
                "You are an expert DuckDB SQL assistant. The user has written the following SQL query, "
                "which may contain comments describing what they want to do.\n"
                "Please fix the query or implement the logic described in the comments.\n"
                "Use DuckDB-specific syntax and functions where appropriate.\n"
                "Return the valid SQL query. You MUST preserve the original comments in the output, "
                "placing them above the corresponding SQL code.\n\n"
                f"Query:\n{sql}"
            )

            # Prepend project context if available
            if active_db:
                project_name = Path(active_db).stem
                prompt = (
                    f"[Context: You are working with the '{project_name}' DuckDB project.]\n"
                    f"{prompt}"
                )

            response = await asyncio.to_thread(agent.run, prompt, reset=False)
            result = self._materialize_agent_response(response).strip()

            # Clean up markdown code blocks if the agent included them despite instructions
            # Remove ```sql or ``` at the start
            result = re.sub(r"^```\w*\s*", "", result)
            # Remove ``` at the end
            result = re.sub(r"\s*```$", "", result)

            return result.strip()

    async def assist_query_error(
        self,
        *,
        sql: str,
        error: str,
        project: Optional[str] = None,
        database: Optional[str] = None,
    ) -> Dict[str, str]:
        """
        Use the AI agent to explain and fix a SQL error.
        """
        agent = await self._ensure_agent()
        async with self._agent_run_lock:
            active_db = self._set_agent_project(agent, project=project, database=database)

            prompt = (
                "You are an expert DuckDB SQL assistant. "
                "The user tried to execute a SQL query but encountered an error.\n"
                "Please explain why the error occurred and provide a corrected SQL query.\n"
                "Use DuckDB-specific syntax and functions where appropriate.\n"
                "Return the result as a JSON object with two keys:\n"
                "1. 'explanation': A concise explanation of the error and the fix.\n"
                "2. 'fixed_sql': The corrected SQL query.\n"
                "Do not include markdown formatting around the JSON.\n\n"
                f"Query:\n{sql}\n\n"
                f"Error:\n{error}"
            )

            # Prepend project context if available
            if active_db:
                project_name = Path(active_db).stem
                prompt = (
                    f"[Context: You are working with the '{project_name}' DuckDB project.]\n"
                    f"{prompt}"
                )

            response = await asyncio.to_thread(agent.run, prompt, reset=False)
            result_text = self._materialize_agent_response(response).strip()

            # Clean up markdown code blocks if present
            result_text = re.sub(r"^```json\s*", "", result_text)
            result_text = re.sub(r"^```\s*", "", result_text)
            result_text = re.sub(r"\s*```$", "", result_text)

            try:
                return json.loads(result_text)
            except json.JSONDecodeError:
                # Fallback if LLM didn't return valid JSON
                return {
                    "explanation": "The AI assistant provided a response but it wasn't in the expected format.",
                    "fixed_sql": result_text
                }

    def _materialize_agent_response(self, response: Any) -> str:
        if response is None:
            return ""

        if isinstance(response, str):
            return response

        if hasattr(response, "output"):
            return self._materialize_agent_response(getattr(response, "output"))

        if isinstance(response, dict):
            for key in ("output", "text", "message", "content"):
                if key in response and response[key]:
                    return self._materialize_agent_response(response[key])
            return str(response)

        if isinstance(response, GeneratorType) or isinstance(response, Iterable):
            fragments: List[str] = []
            for item in response:
                text = self._materialize_agent_response(item)
                if text:
                    fragments.append(text)
            return "".join(fragments)

        return str(response)

    def _serialize_tool_calls(self, tool_calls: Optional[List[Any]]) -> Optional[List[Dict[str, Any]]]:
        if not tool_calls:
            return None
        serialized = []
        for tc in tool_calls:
            if isinstance(tc, dict):
                serialized.append(tc)
            elif hasattr(tc, "model_dump"):
                serialized.append(tc.model_dump())
            elif hasattr(tc, "dict"):
                serialized.append(tc.dict())
            else:
                try:
                    serialized.append(dict(tc))
                except (ValueError, TypeError):
                    pass
        return serialized

    # Project metadata helpers -------------------------------------------------

    def _project_metadata_from_path(self, path: Path) -> Dict[str, Any]:
        schema_path = path.with_suffix(".schema.json")
        display_name = path.stem
        description = ""
        version = None
        subdirectory = ""

        # Check if file is in a subdirectory relative to search roots
        # Use relative path as unique name identifier
        unique_name = path.stem
        for root in self.query_tool.search_roots or [Path.cwd()]:
            try:
                rel = path.relative_to(root)
                if len(rel.parts) > 1:  # Has subdirectory
                    subdirectory = str(rel.parent)
                    # Use relative path without extension as unique name
                    unique_name = str(rel.with_suffix("")).replace("\\", "/")
                break
            except ValueError:
                continue

        if schema_path.exists():
            try:
                payload = json.loads(schema_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                payload = {}
            display_name = str(payload.get("project_display_name") or payload.get("project") or path.stem)
            description = str(payload.get("project_description") or "")
            version = payload.get("version")
            query_instructions = str(payload.get("query_instructions") or "")
        else:
            query_instructions = ""

        return {
            "name": unique_name,
            "path": str(path),
            "display_name": display_name,
            "subdirectory": subdirectory,
            "description": description,
            "version": version,
            "query_instructions": query_instructions,
        }

    def _slugify(self, value: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("_")
        return cleaned.lower()

    async def update_project_metadata(
        self,
        *,
        path: str,
        display_name: Optional[str] = None,
        description: Optional[str] = None,
        version: Optional[str] = None,
        query_instructions: Optional[str] = None,
    ) -> Dict[str, Any]:
        target = Path(path).expanduser().resolve()
        if not target.exists():
            raise ValueError("The specified project database does not exist.")

        async with self._query_lock:
            manager = SchemaManager()
            manager.set_database(target)
            current_meta = manager.get_project_metadata()

            # Update metadata first (persists to current schema file)
            manager.set_project_metadata(
                display_name=display_name if display_name is not None else current_meta["display_name"],
                description=description if description is not None else current_meta["description"],
                version=version if version is not None else current_meta.get("version"),
                query_instructions=query_instructions
                if query_instructions is not None
                else current_meta.get("query_instructions"),
            )

            # Handle renaming if version changed
            if version and version != current_meta.get("version"):
                # Determine base name
                stem = target.stem
                # Regex to match _vX.Y.Z at the end
                match = re.search(r"^(.*)_v\d+(\.\d+)*$", stem)
                if match:
                    base_name = match.group(1)
                else:
                    base_name = stem

                new_stem = f"{base_name}_v{version}"
                new_target = target.with_name(f"{new_stem}.duckdb")

                # Rename .duckdb file
                if new_target != target:
                    try:
                        target.rename(new_target)

                        # Rename .schema.json file
                        old_schema = target.with_suffix(".schema.json")
                        new_schema = new_target.with_suffix(".schema.json")
                        if old_schema.exists():
                            old_schema.rename(new_schema)

                        # Update active database reference if needed
                        # We compare against the original resolved target
                        if self.query_tool.current_database == target:
                            self.query_tool.current_database = new_target

                        # Update target to new path for return value
                        target = new_target

                        # Update schema manager's path to ensure consistency
                        manager.set_database(target)
                    except OSError as e:
                        print(f"Failed to rename project files: {e}")
                        # We don't raise here to at least return the metadata update,
                        # but the file remains with old name.

            info = self._project_metadata_from_path(target)
            current = self.query_tool.current_database
            info["is_active"] = bool(current and target == current)
            return info

    async def create_project(
        self, *, name: str, description: Optional[str] = "", query_instructions: Optional[str] = ""
    ) -> Dict[str, Any]:
        trimmed = name.strip()
        if not trimmed:
            raise ValueError("Project name must not be empty.")

        slug = self._slugify(trimmed)
        if not slug:
            raise ValueError("Project name must contain alphanumeric characters.")

        search_roots = self.query_tool.search_roots or [Path.cwd()]
        root = search_roots[0]
        root.mkdir(parents=True, exist_ok=True)

        candidate = root / f"{slug}.duckdb"
        counter = 1
        while candidate.exists():
            candidate = root / f"{slug}_{counter}.duckdb"
            counter += 1

        await asyncio.to_thread(self._initialise_duckdb, candidate)

        manager = SchemaManager()
        manager.set_database(candidate)
        manager.set_project_metadata(
            display_name=trimmed, description=description or "", query_instructions=query_instructions or ""
        )

        info = self._project_metadata_from_path(candidate)
        info["is_active"] = False
        return info

    @staticmethod
    def _initialise_duckdb(path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with duckdb.connect(str(path)) as conn:
            conn.execute("SELECT 1")

    # Datalake sync methods -------------------------------------------

    async def list_datalakes(self) -> List[Dict[str, str]]:
        """List all configured datalakes."""
        if not self.datalake_manager:
            return []

        def get_all_info():
            result = []
            if self.datalake_manager:
                for name in self.datalake_manager.list_datalakes():
                    info = self.datalake_manager.get_datalake_info(name)
                    if info:
                        result.append(info)
            return result

        return await asyncio.to_thread(get_all_info)

    async def list_datalake_projects(self, datalake_name: str) -> List[Dict[str, Any]]:
        """List projects in a specific datalake."""
        if not self.datalake_manager:
            raise ValueError("No datalakes configured")

        projects = await asyncio.to_thread(
            self.datalake_manager.list_projects,
            datalake_name
        )

        return [
            {
                "name": p.name,
                "version": p.version,
                "display_name": p.display_name,
                "description": p.description,
                "last_modified": p.last_modified,
                "size_bytes": p.size_bytes,
                "db_size_bytes": p.db_size_bytes,
                "schema_size_bytes": p.schema_size_bytes,
            }
            for p in projects
        ]

    async def download_project_from_datalake(
        self,
        datalake_name: str,
        project_name: str,
        version: str,
        overwrite: bool = False,
        rename_existing: bool = False,
    ) -> Dict[str, Any]:
        """Download a project from a datalake."""
        if not self.datalake_manager:
            raise ValueError("No datalakes configured")

        search_roots = self.query_tool.search_roots or [Path.cwd()]
        local_dir = search_roots[0]

        # If rename_existing is True and files exist, rename them with version
        if rename_existing and not overwrite:
            duckdb_file = local_dir / f"{project_name}.duckdb"
            schema_file = local_dir / f"{project_name}.schema.json"

            if duckdb_file.exists() or schema_file.exists():
                # Get current version from schema
                current_version = "unknown"
                if schema_file.exists():
                    try:
                        schema_data = json.loads(schema_file.read_text(encoding="utf-8"))
                        current_version = schema_data.get("version", "unknown")
                    except (json.JSONDecodeError, OSError):
                        pass

                # Rename existing files
                if duckdb_file.exists():
                    new_name = f"{project_name}_v{current_version}.duckdb"
                    duckdb_file.rename(local_dir / new_name)

                if schema_file.exists():
                    new_name = f"{project_name}_v{current_version}.schema.json"
                    schema_file.rename(local_dir / new_name)

        # Download the project
        duckdb_path, schema_path = await asyncio.to_thread(
            self.datalake_manager.download_project,
            datalake_name,
            project_name,
            version,
            local_dir,
            overwrite=overwrite,
        )

        return {
            "success": True,
            "project_name": project_name,
            "version": version,
            "duckdb_path": str(duckdb_path),
            "schema_path": str(schema_path),
            "message": f"Downloaded project '{project_name}' version {version}",
        }

    async def upload_project_to_datalake(
        self,
        datalake_name: str,
        project_path: str,
        new_version: Optional[str] = None,
        schema_only: bool = False,
    ) -> Dict[str, Any]:
        """Upload a project to a datalake."""
        if not self.datalake_manager:
            raise ValueError("No datalakes configured")

        duckdb_path = Path(project_path)
        if not duckdb_path.exists():
            raise ValueError(f"Project file not found: {project_path}")

        schema_path = duckdb_path.parent / f"{duckdb_path.stem}.schema.json"
        if not schema_path.exists():
            raise ValueError(f"Schema file not found: {schema_path}")

        # Get current version from schema
        schema_data = json.loads(schema_path.read_text(encoding="utf-8"))
        current_version = schema_data.get("version", "1.0.0")

        # Determine version to upload
        if new_version:
            version = new_version
            # Update version in schema
            schema_data["version"] = version
            schema_path.write_text(json.dumps(schema_data, indent=2), encoding="utf-8")
        else:
            version = current_version

        # Determine project name by stripping version suffix if present
        # This ensures we upload to dbdocumenter/project/version/ instead of dbdocumenter/project_vX/version/
        project_name = duckdb_path.stem
        match = re.search(r"^(.*)_v\d+(\.\d+)*$", project_name)
        if match:
            project_name = match.group(1)

        # Upload the project
        await asyncio.to_thread(
            self.datalake_manager.upload_project,
            datalake_name,
            project_name,
            version,
            duckdb_path,
            schema_path,
            schema_only=schema_only,
        )

        return {
            "success": True,
            "project_name": project_name,
            "version": version,
            "message": f"Uploaded project '{project_name}' version {version}",
        }

    async def add_datalake(
        self,
        name: str,
        connection_string: str,
        type: str = "azure_storage",
        container_name: str = "dbdocumenter",
    ) -> Dict[str, str]:
        """Add a new datalake configuration (runtime only, not persisted)."""
        if not self.datalake_manager:
            from ..tools.datalake.manager import DatalakeManager as DLManager

            self.datalake_manager = DLManager([])

        from ..tools.datalake.manager import DatalakeConfig

        config = DatalakeConfig(
            name=name,
            type=type,
            connection_string=connection_string,
            container_name=container_name,
        )

        await asyncio.to_thread(self.datalake_manager.add_datalake, config, False)

        # Persist to config file
        await self._persist_datalakes()

        # Extract storage account name from connection string
        storage_account = "Unknown"
        if type == "azure_storage":
            try:
                parts = connection_string.split(";")
                for part in parts:
                    if part.startswith("AccountName="):
                        storage_account = part.split("=", 1)[1]
                        break
            except Exception:
                pass

        return {
            "name": name,
            "type": type,
            "container_name": container_name,
            "storage_account": storage_account,
        }

    async def test_datalake_connection(
        self,
        type: str,
        connection_string: str,
    ) -> Dict[str, Any]:
        """Test datalake connection and list available containers."""
        if type != "azure_storage":
            return {
                "success": False,
                "message": f"Unsupported datalake type: {type}",
                "containers": [],
            }

        try:
            from azure.storage.blob import BlobServiceClient

            # Test connection by attempting to create a BlobServiceClient
            blob_service = BlobServiceClient.from_connection_string(connection_string)

            # List containers
            containers = []
            for container in blob_service.list_containers():
                containers.append(container.name)

            return {
                "success": True,
                "message": f"Successfully connected. Found {len(containers)} container(s).",
                "containers": containers,
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"Connection failed: {str(e)}",
                "containers": [],
            }

    async def remove_datalake(self, name: str) -> bool:
        """Remove a datalake configuration (runtime only)."""
        if not self.datalake_manager:
            return False

        result = await asyncio.to_thread(self.datalake_manager.remove_datalake, name)

        if result:
            # Persist to config file
            await self._persist_datalakes()

        return result

    async def _persist_datalakes(self) -> None:
        """Persist runtime-added datalakes to config file."""
        if not self.datalake_manager:
            return

        from ..server.config import DatalakeConfig as ConfigDatalakeConfig
        from ..server.config import save_datalakes_config

        def save_configs():
            if self.datalake_manager:
                persistable = self.datalake_manager.get_persistable_configs()
                # Convert to config module's DatalakeConfig
                configs = [
                    ConfigDatalakeConfig(
                        name=c.name,
                        type=c.type,
                        connection_string=c.connection_string,
                        container_name=c.container_name,
                    )
                    for c in persistable
                ]
                save_datalakes_config(configs)

        await asyncio.to_thread(save_configs)

    async def delete_datalake_project(
        self,
        datalake_name: str,
        project_name: str,
        version: str,
    ) -> bool:
        """Delete a project version from a datalake."""
        if not self.datalake_manager:
            raise ValueError("No datalakes configured")

        await asyncio.to_thread(
            self.datalake_manager.delete_project,
            datalake_name,
            project_name,
            version,
        )

        return True

    async def delete_field(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
        field: str,
    ) -> None:
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available to delete field.")
            manager = self.schema_manager
            removed = manager.delete_field(table, field)
            if not removed:
                raise ValueError(f"Field '{field}' not found in table '{table}'.")

    async def enrich_table_from_file(
        self,
        *,
        project: Optional[str],
        database: Optional[str],
        table: str,
        filename: str,
        file_content: bytes,
    ) -> Dict[str, Any]:
        """
        Enrich table schema using metadata from an uploaded file (CSV, Excel, TXT).
        """
        async with self._query_lock:
            resolved = await self._ensure_project_locked(project=project, database=database)
            if resolved is None:
                raise ValueError("No DuckDB database available.")

            # 1. Parse the file
            import io
            content_text = ""
            ext = Path(filename).suffix.lower()

            try:
                if ext == ".csv":
                    import csv

                    # Decode bytes to string
                    text_content = file_content.decode("utf-8", errors="ignore")
                    f = io.StringIO(text_content)
                    reader = csv.reader(f)
                    rows = []
                    for i, row in enumerate(reader):
                        if i >= 20:
                            break
                        rows.append(row)

                    if rows:
                        headers = rows[0]
                        md_lines = ["| " + " | ".join(str(h) for h in headers) + " |"]
                        md_lines.append("| " + " | ".join("---" for _ in headers) + " |")
                        for row in rows[1:]:
                            md_lines.append("| " + " | ".join(str(c) for c in row) + " |")
                        content_text = "\n".join(md_lines)

                elif ext == ".xlsx":
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                        ws = wb.active
                        rows = []
                        if ws:
                            for i, row in enumerate(ws.iter_rows(values_only=True)):
                                if i >= 20:
                                    break
                                rows.append(list(row))

                        if rows:
                            headers = rows[0]
                            md_lines = ["| " + " | ".join(str(h) for h in headers) + " |"]
                            md_lines.append("| " + " | ".join("---" for _ in headers) + " |")
                            for row in rows[1:]:
                                md_lines.append("| " + " | ".join(str(c) if c is not None else "" for c in row) + " |")
                            content_text = "\n".join(md_lines)
                    except ImportError:
                        raise RuntimeError("openpyxl is required for Excel files.")

                elif ext == ".xls":
                    raise ValueError(".xls files are not supported without pandas. Please convert to .xlsx or .csv")
                else:
                    # Assume text
                    content_text = file_content.decode("utf-8", errors="ignore")
                    # Limit length
                    if len(content_text) > 10000:
                        content_text = content_text[:10000] + "...(truncated)"
            except Exception as e:
                raise ValueError(f"Failed to parse file '{filename}': {str(e)}")

            # 2. Get current schema
            manager = self.schema_manager
            table_record = manager.get_table(table)
            if not table_record:
                raise ValueError(f"Table '{table}' not found in schema.")

            # Get documented fields
            documented_fields = set(table_record.get("fields", {}).keys())

            # Get actual database fields to include undocumented ones
            db_fields = set()
            try:
                schema_rows = await asyncio.to_thread(self.query_tool.fetch_table_schema, table)
                for row in schema_rows:
                    # PRAGMA table_info returns 'name' column
                    if "name" in row:
                        db_fields.add(str(row["name"]))
            except Exception:
                pass

            # Merge fields (documented + actual)
            fields = list(documented_fields.union(db_fields))

            if not fields:
                raise ValueError(f"No fields found for table '{table}'.")

            # 3. Ask AI to map
            agent = await self._ensure_agent()

            prompt = (
                "You are a schema enrichment assistant. Your task is to extract field descriptions "
                "from a provided file content and map them to the fields of a specific database table.\n\n"
                f"Target Table: {table}\n"
                f"Target Fields: {', '.join(fields)}\n\n"
                "Input File Content:\n"
                f"{content_text}\n\n"
                "Instructions:\n"
                "1. Analyze the Input File Content to find descriptions or definitions that match the Target Fields.\n"
                "2. Match fields based on name similarity or semantic meaning.\n"
                "3. Return a JSON object where keys are the exact Target Field names "
                "and values are the extracted descriptions.\n"
                "4. Only include fields where you found a confident match.\n"
                "5. Return ONLY the valid JSON string. Do not include markdown formatting "
                "(like ```json) or explanations.\n"
            )

            response = await asyncio.to_thread(agent.run, prompt, reset=False)
            result_text = self._materialize_agent_response(response).strip()

            # Clean up markdown if present
            result_text = re.sub(r"^```\w*\s*", "", result_text)
            result_text = re.sub(r"\s*```$", "", result_text)

            try:
                mapping = json.loads(result_text)
            except json.JSONDecodeError:
                raise ValueError("AI failed to return valid JSON mapping.")

            # 4. Apply updates
            updated_count = 0
            for field, description in mapping.items():
                if field in fields and description:
                    # We assume these are long descriptions if they are detailed,
                    # or we could try to split them. For now, let's put it in long_description
                    # and generate a short one if missing?
                    # Or just put it in long_description.
                    # The user asked to "enrich the fields".

                    # Let's put it in long_description as it's safer for "enrichment"
                    manager.update_field(table, field, long_description=description)

                    # If short description is missing, maybe use the first sentence?
                    # For now, let's just update long_description.
                    updated_count += 1

            return {
                "success": True,
                "table": table,
                "updated_fields": updated_count,
                "mapping": mapping
            }

    def _process_file_content(self, filename: str, content: str) -> str:
        """
        Process uploaded file content.
        If it's a base64 data URL (Excel), decode and parse with openpyxl.
        Otherwise return as is.
        """
        if content.startswith("data:") and ";base64," in content:
            try:
                import base64
                import io

                _, encoded = content.split(";base64,", 1)
                decoded = base64.b64decode(encoded)

                ext = Path(filename).suffix.lower()
                if ext == ".xlsx":
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(decoded), read_only=True, data_only=True)
                        ws = wb.active
                        rows = []
                        if ws:
                            for i, row in enumerate(ws.iter_rows(values_only=True)):
                                if i >= 50:
                                    break
                                rows.append(list(row))

                        if not rows:
                            return "[Empty Excel file]"

                        # Simple markdown table generation
                        headers = rows[0]
                        md_lines = []
                        # Header
                        md_lines.append("| " + " | ".join(str(h) for h in headers) + " |")
                        # Separator
                        md_lines.append("| " + " | ".join("---" for _ in headers) + " |")
                        # Rows
                        for row in rows[1:]:
                            md_lines.append("| " + " | ".join(str(c) if c is not None else "" for c in row) + " |")

                        return "\n".join(md_lines)
                    except ImportError:
                        return "[Error: openpyxl is required to process Excel files. Please install it.]"
                elif ext == ".xls":
                    return "[Error: .xls files are not supported without pandas. Please convert to .xlsx or .csv]"
                else:
                    # Try to decode as utf-8 text if not excel
                    try:
                        return decoded.decode("utf-8")
                    except UnicodeDecodeError:
                        return f"[Binary file {filename} uploaded. Size: {len(decoded)} bytes. Content not displayed.]"
            except Exception as e:
                return f"[Error processing file {filename}: {str(e)}]"

        return content

    def _sanitize_history(self, history: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Ensure history is valid for OpenAI API.
        Specifically, ensure every tool call has a corresponding tool response.
        """
        sanitized = []
        pending_tool_calls = {}  # id -> tool_call_dict

        for msg in history:
            role = msg.get("role")

            # If we have pending tool calls and we see a non-tool message,
            if pending_tool_calls and role != "tool":
                # This shouldn't happen in a valid conversation, but if it does,
                # we need to close out the pending calls.
                for tool_call_id in list(pending_tool_calls.keys()):
                    sanitized.append({
                        "role": "tool",
                        "tool_call_id": tool_call_id,
                        "content": (
                            "Error: Tool execution was interrupted or response was lost."
                        )
                    })
                    del pending_tool_calls[tool_call_id]

            sanitized.append(msg)

            if role == "assistant" and msg.get("tool_calls"):
                for tc in msg["tool_calls"]:
                    # Handle both object and dict access
                    tc_id = tc.get("id") if isinstance(tc, dict) else getattr(tc, "id", None)
                    if tc_id:
                        pending_tool_calls[tc_id] = tc

            elif role == "tool":
                tc_id = msg.get("tool_call_id")
                if tc_id in pending_tool_calls:
                    del pending_tool_calls[tc_id]

        # If we reached the end and still have pending tool calls, close them out
        for tool_call_id in pending_tool_calls:
            sanitized.append({
                "role": "tool",
                "tool_call_id": tool_call_id,
                "content": (
                    "Error: Tool execution was interrupted or response was lost."
                )
            })

        return sanitized
